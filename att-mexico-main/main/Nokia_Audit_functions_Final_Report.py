def final_report(sesion_id_t, consecutivo_t, fileLocation, Logs_Path):

    print("Rep FINAL: ",sesion_id_t, consecutivo_t, fileLocation, Logs_Path)
    # In[1]:


    #funciones
    from openpyxl import load_workbook
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows
    import os
    import pandas as pd
    import openpyxl as op


    # In[12]:


    #Datos de sesion - seran reemplazados cuando se pase a produccion
    def datos_de_sesion(sesion_id_t,consecutivo_t, root_dir):


        sesion_id=sesion_id_t
        file_dir=root_dir+sesion_id               #file_dir directorio + sesion donde estan alojados los archivos
        consecutivo=consecutivo_t+'.csv'
        print("file_dir:",file_dir)
        return(file_dir, consecutivo)


    #Función lista_imagenes : Listar carpeta de sesion y retornar lista de imagenes

    def lista_imagenes(input_path):                    #input_path contiene la ruta de la sesion
        file_names= os.listdir(input_path)             #lista de archivos en carpeta
        image_names=[]
        for nombre in file_names:                      #recorrer lista
            if nombre.split(".")[-1] not in ["png"]:   #seleccionar extension png
                continue
            imagen_path= file_dir+"/"+nombre
            image_names.append(imagen_path)            #Lista con directorio de imagenes -> para guardar en excel
        return(image_names)



    #Funcion obtener listado de coordenadas

    def lista_coord(png):
        celdas=[]
        v_letras=['A','L','W']               #columnas indice
        fotos=1
        col=0
        fil=1
        while( fotos<= len(png)):
            coor=v_letras[col]+str(fil)
            celdas.append(coor)
            if(v_letras[col]=='W'):          # cambiar de fila
                col=0
                fil=fil+31
            else:
                col=col+1                    #recorrer 3 columnas
            fotos=fotos+1                    #recorrer lista fotos
        return celdas


    def ingesta_mapas(fileLocation,writeLocation,lista_png,coordenadas):
        book = load_workbook(fileLocation)          #cargar en memoria template
        sheet = book.create_sheet("Distribución de TAC LAC") #crear nueva pestaña en memoria
        try:
            for i in range(0,len(lista_png)):           # llenado de imagenes 
                imag = Image(lista_png[i])
                sheet.add_image(imag, coordenadas[i])
        except:
            pass
        book.save(writeLocation)                    #guardado de excel
        return



    def ingesta_audit(source_raw,pestana_activa):
        filesource_name= file_dir+"/"+source_raw+consecutivo
        df=pd.read_csv(filesource_name)
        df=pd.DataFrame(df)   #DF origen
        df=df.drop(columns=['Unnamed: 0'])
        book = load_workbook(fileLocation)                             #cargar en memoria actual libro
        sheet = book[pestana_activa]                                   #cargar en memoria pestaña destino
        for row in dataframe_to_rows(df, index=False, header=False):   #llenado df en pestaña
            sheet.append(row)
        book.save(writeLocation)
        print(pestana_activa+": ejecutada")
        print(filesource_name)
        return


    #incrustar excel detalle como objeto en entregable. Este paso debe ser el ultimo.





    #incia=datos_de_sesion('20H-34M-40S_59711','20H-35M-37S_2010.csv')

    file_dir, consecutivo=datos_de_sesion(sesion_id_t, consecutivo_t, Logs_Path)

    # In[3]:


    lista_png =lista_imagenes(file_dir)
    #lista_png


    # In[4]:


    coordenadas= lista_coord(lista_png)
    #coordenadas


    # In[5]:


    #Funcion Cargar imagenes en archivo existente nueva pestaña

    #fileLocation ='D:\ClickHouse_NOKIA\ATT_Visual\Template_Audit.xlsx' #ubicacion del template

    writeLocation=file_dir+'/Audit_Report_FINAL.xlsx'



    mapas=ingesta_mapas(fileLocation,writeLocation,lista_png,coordenadas)
    print(writeLocation)


    # In[6]:


    #Adaptar KPIS de .csv a .xlsx
    filesource_kpi= file_dir+"/"+"KPIS_LTE_Log_"+consecutivo
    filesource_kpi_csv = filesource_kpi#.replace('/','\\\\')
    filesource_kpi_xlsx =filesource_kpi_csv.replace(".csv",".xlsx")

    read_file = pd.read_csv (filesource_kpi_csv,header=None)
    #print(read_file)
    read_file.to_excel (filesource_kpi_xlsx, index = None, header=None)
    #print(read_file)


    # In[7]:

    #Cargar KPIS desde excel y guardarlos en entregable. fUNCION REQUIERE \\

    wb1 = op.load_workbook(filesource_kpi_xlsx) 
    ws1 = wb1.worksheets[0] 

    filename1 =writeLocation#.replace('/','\\\\')
    wb2 = op.load_workbook(filename1) 
    ws2 = wb2.active 
      
    mr = ws1.max_row                            #MAX Dimension fila
    mc = ws1.max_column                         #MAX Dimension columna

    print("MR",mr)
    print("MC",mc)
    fila_text= [2,3,4]                          #Dejar filas como texto
    for i in range (1, mr + 1):                 #iteracion fila
        for j in range (1, mc + 1):                   #iteracion columna
            c = ws1.cell(row = i, column = j) 
            #print("Value-->",c)
            if i in fila_text or j==1 : 
                ws2.cell(row = i+2, column = j+2).value = c.value      #ubicacion del cursor +2 , +2 celda C3
            else:
                ws2.cell(row = i+2, column = j+2).value = float(c.value)    #convertir a valor
    wb2.save(str(filename1)) 
      


    # In[8]:


    #funcion lectura data plana

    fileLocation =writeLocation                             #editar sobre el entregable


    # In[11]:
    reports_list=[]

    reports_list.append(['Audit_RET_Baseline_Diff_Log_','Diferencia RET vs TE'])
    reports_list.append(['Audit_RET_Same_Sector_Log_','RET del mismo Sector'])
    reports_list.append(['Audit_RSSI_Th_Detail_Log_','Umbral RSSI'])
    reports_list.append(['Audit_RTWP_Diff_Detail_Log_','Diferencia entre puertos RTWP'])
    reports_list.append(['Audit_RTWP_Th_Detail_Log_','Umbral RTWP'])
    reports_list.append(['LNBTS_HW_Audit_Log_','LNBTS HW'])
    reports_list.append(['PCI_Conflicts_Detail_Log_','Conflictos PCI'])
    reports_list.append(['RSI_Conflicts_Detail_Log_','Conflictos RSI'])
    reports_list.append(['Param_Audit_Summary_Log_','Resumen Parametros'])
    reports_list.append(['Audit_LNRELW_Summary_Log_','Resumen LNRELW'])
    reports_list.append(['PCI_Conflicts_Detail_Log_','Conflictos PCI'])

    for report in reports_list:
        try:
            a=ingesta_audit(report[0],report[1])
        except:
            print("Fallo Reporte:",report[1])







def final_report_kpi(sesion_id_t, consecutivo_t, fileLocation, Logs_Path):

    # In[1]:


    #funciones
    from openpyxl import load_workbook
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows
    import os
    import pandas as pd
    import openpyxl as op


    # In[12]:


    #Datos de sesion - seran reemplazados cuando se pase a produccion
    def datos_de_sesion(sesion_id_t,consecutivo_t, root_dir):


        sesion_id=sesion_id_t

        file_dir=root_dir+sesion_id               #file_dir directorio + sesion donde estan alojados los archivos
        consecutivo=consecutivo_t+'.csv'
        print(file_dir)
        return(file_dir, consecutivo)




    def ingesta_audit(source_raw,pestana_activa):
        filesource_name= file_dir+"/"+source_raw+consecutivo
        df=pd.read_csv(filesource_name)
        df=pd.DataFrame(df)   #DF origen
        df=df.drop(columns=['Unnamed: 0'])
        book = load_workbook(fileLocation)                             #cargar en memoria actual libro
        sheet = book[pestana_activa]                                   #cargar en memoria pestaña destino
        for row in dataframe_to_rows(df, index=False, header=False):   #llenado df en pestaña
            sheet.append(row)
        book.save(writeLocation)
        print(pestana_activa+": ejecutada")
        print(filesource_name)
        return

    def ingesta_ini(fileLocation,writeLocation):
        book = load_workbook(fileLocation)          #cargar en memoria template
        sheet = book.create_sheet("Distribución de TAC LAC") #crear nueva pestaña en memoria
        try:
            for i in range(0,len(lista_png)):           # llenado de imagenes 
                imag = Image(lista_png[i])
                sheet.add_image(imag, coordenadas[i])
        except:
            pass
        book.save(writeLocation)                    #guardado de excel
        return



    file_dir, consecutivo=datos_de_sesion(sesion_id_t, consecutivo_t, Logs_Path)


    # In[5]:


    #Funcion Cargar imagenes en archivo existente nueva pestaña



    writeLocation=file_dir+'/Audit_Report_FINAL.xlsx'


    # In[6]:
    mapas=ingesta_ini(fileLocation,writeLocation)
    print(writeLocation)

    #Adaptar KPIS de .csv a .xlsx
    filesource_kpi= file_dir+"/"+"KPIS_LTE_Log_"+consecutivo
    filesource_kpi_csv = filesource_kpi#.replace('/','\\\\')
    filesource_kpi_xlsx =filesource_kpi_csv.replace(".csv",".xlsx")

    read_file = pd.read_csv (filesource_kpi_csv,header=None)
    #display(read_file)
    read_file.to_excel (filesource_kpi_xlsx, index = None, header=None)
    #display(read_file)


    # In[7]:

    #Cargar KPIS desde excel y guardarlos en entregable. fUNCION REQUIERE \\

    wb1 = op.load_workbook(filesource_kpi_xlsx) 
    ws1 = wb1.worksheets[0] 

    filename1 =writeLocation#.replace('/','\\\\')
    wb2 = op.load_workbook(filename1) 
    ws2 = wb2.active 
      
    mr = ws1.max_row                            #MAX Dimension fila
    mc = ws1.max_column                         #MAX Dimension columna
    fila_text= [2,3,4]                          #Dejar filas como texto
    for i in range (1, mr + 1):                 #iteracion fila
        for j in range (1, mc + 1):                   #iteracion columna
            c = ws1.cell(row = i, column = j) 
            if i in fila_text or j==1 : 
                ws2.cell(row = i+2, column = j+2).value = c.value      #ubicacion del cursor +2 , +2 celda C3
            else:
                ws2.cell(row = i+2, column = j+2).value = float(c.value)    #convertir a valor
    wb2.save(str(filename1)) 
      


    # In[8]:


    #funcion lectura data plana

    fileLocation =writeLocation                             #editar sobre el entregable


